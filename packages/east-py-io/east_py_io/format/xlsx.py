"""XLSX platform functions for East.

Provides Excel file reading and writing for East programs.
"""

import io
from datetime import datetime
from typing import Any

from east.runtime.platform import PlatformFunction
from east.types.types import BlobType
from east.types.values import EastArray, EastBlob, EastStruct, EastVariant
from openpyxl import Workbook, load_workbook

from .types import (
    LiteralValueType,
    XlsxInfoType,
    XlsxReadOptionsType,
    XlsxRowType,
    XlsxSheetInfoType,
    XlsxSheetType,
    XlsxWriteOptionsType,
)


def convert_cell_to_east(value: Any) -> EastVariant:
    """Convert an Excel cell value to East LiteralValueType variant."""
    if value is None:
        return EastVariant("Null", None)
    elif isinstance(value, bool):
        return EastVariant("Boolean", value)
    elif isinstance(value, int):
        return EastVariant("Integer", value)
    elif isinstance(value, float):
        return EastVariant("Float", value)
    elif isinstance(value, datetime):
        return EastVariant("DateTime", value)
    elif isinstance(value, str):
        return EastVariant("String", value)
    else:
        return EastVariant("String", str(value))


def convert_east_to_cell(value: EastVariant) -> Any:
    """Convert East LiteralValueType variant to Excel cell value."""
    tag = value.type
    val = value.value

    if tag == "Null":
        return None
    elif tag == "Boolean":
        return val
    elif tag == "Integer":
        return int(val) if val is not None else 0
    elif tag == "DateTime":
        return val
    elif tag == "Blob":
        return val.hex() if hasattr(val, "hex") else str(val)
    else:
        return val


async def xlsx_read_impl(blob: EastBlob, options: EastStruct) -> EastArray:
    """Read an XLSX file."""
    try:
        # Get options
        sheet_name_opt = options["sheetName"]
        sheet_name = sheet_name_opt.value if sheet_name_opt.type == "some" else None

        # Load workbook
        wb = load_workbook(filename=io.BytesIO(bytes(blob)), read_only=True, data_only=True)

        # Get sheet
        ws = wb[sheet_name] if sheet_name else wb.active

        if ws is None:
            wb.close()
            return EastArray(XlsxRowType, [])

        # Read data
        result = EastArray(XlsxRowType, [])
        for row in ws.iter_rows():
            row_data = EastArray(
                LiteralValueType, [convert_cell_to_east(cell.value) for cell in row]
            )
            result.append(row_data)

        wb.close()
        return result
    except Exception as e:
        raise Exception(f"XLSX read failed: {e}") from e


async def xlsx_write_impl(data: EastArray, options: EastStruct) -> EastBlob:
    """Write data to an XLSX file."""
    try:
        # Get options
        sheet_name_opt = options["sheetName"]
        sheet_name = sheet_name_opt.value if sheet_name_opt.type == "some" else "Sheet1"

        # Create workbook
        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet()
        ws.title = sheet_name

        # Write data
        for row_idx, row in enumerate(data, start=1):
            for col_idx, cell in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=convert_east_to_cell(cell))

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        wb.close()

        return EastBlob(output.getvalue())
    except Exception as e:
        raise Exception(f"XLSX write failed: {e}") from e


async def xlsx_info_impl(blob: EastBlob) -> EastStruct:
    """Get information about an XLSX file."""
    try:
        wb = load_workbook(filename=io.BytesIO(bytes(blob)), read_only=True)

        sheets = EastArray(XlsxSheetInfoType, [])
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheets.append(
                EastStruct(
                    {
                        "name": sheet_name,
                        "rowCount": ws.max_row or 0,
                        "columnCount": ws.max_column or 0,
                    }
                )
            )

        wb.close()
        return EastStruct({"sheets": sheets})
    except Exception as e:
        raise Exception(f"XLSX info failed: {e}") from e


# Platform function implementations
xlsx_impl = [
    PlatformFunction(
        name="xlsx_read",
        inputs=[BlobType, XlsxReadOptionsType],
        output=XlsxSheetType,
        type="async",
        fn=xlsx_read_impl,
    ),
    PlatformFunction(
        name="xlsx_write",
        inputs=[XlsxSheetType, XlsxWriteOptionsType],
        output=BlobType,
        type="async",
        fn=xlsx_write_impl,
    ),
    PlatformFunction(
        name="xlsx_info",
        inputs=[BlobType],
        output=XlsxInfoType,
        type="async",
        fn=xlsx_info_impl,
    ),
]

__all__ = [
    "xlsx_impl",
    "xlsx_read_impl",
    "xlsx_write_impl",
    "xlsx_info_impl",
]
